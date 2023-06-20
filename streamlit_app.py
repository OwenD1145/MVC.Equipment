import pandas as pd
import streamlit as st 
from PIL import Image
from openpyxl import Workbook, load_workbook
import io
from copy import copy
 
st.set_page_config(
  page_title="Equipment Dashboard",
  page_icon=":bar_chart:",
  layout="wide"                 
)

# MAINPAGE

st.header(":space_invader: MVC Equipment Schedule Automation")
st.markdown("##")

if st.button('Read Me'):
    st.text('Please add your personal downloads folder to Excels trusted locations to enable Macros in the produced document:') 
    st.code('File -> Options -> Trust Center -> Trust Center Settings -> Trusted Locations -> Add New Location')
    st.text('The file path will be:')
    st.code('C:/Users/user_directory/Downloads')  

st.text('Loads must be seperated by a space between the value and the load type. Horsepower must be written out in fractional form. MOP/MOCP must be listed first. Table below for example formatting:')
ex = pd.DataFrame(
    {
        'Key': ['CP', '1', 'EF', '2', 'FCU', '1'],
        'Equipment': ['Circulation Pump', '', 'Exhaust Fan', '', 'Fan Coil Unit', ''],
        'Load': ['1-1/2 HP', '', '3.5 FLA', '','25 MOCP', '19 MCA' ],
        'Volts': ['120', '', '120', '', '208',''],
        '%%C': ['1', '', '1', '', '1', ''], 
        'Conductors' : ['', '', '', '', '', '',],
        'Conduit' : ['', '', '', '', '', '',],
        'Switch' : ['', '', '', '', '', '',],
        'Fuse' : ['', '', '', '', '', '',],
        }
)

st.dataframe(width = 1500,data = ex)
 
uploaded_file = st.file_uploader("Supply Your Schedule: Must be .XLSM file type (if working with older .XLS file please send to Owen)")
     
def display_xl():
  if uploaded_file is not None:
    
    df = pd.read_excel(
    uploaded_file,
    # io= uploaded_file.name,
    engine='openpyxl',
    sheet_name='Equipment',
    skiprows=1,
    usecols='A:K',
    dtype = {'VOLTS' : str, '%%C' : str, 'NOTES' : str, 'KEY' : str}
    
    # nrows=10,    
    )
  return df
 
def Automation_xl(uploaded_file):
  global WB  
  WB = load_workbook(uploaded_file, data_only = True, read_only = False, keep_vba = True)
  WS = WB.active
  SH = WB["Equipment"]
  
  table_electric_heaters_120_1_KW = {
      '0.5' : ['2#12,1#12G.', '1/2"', '$SPST', '---'],
      '1.5' : ['2#12,1#12G.', '1/2"', '$SPST', '---'],
  }
  
  table_electric_heaters_208_1_KW = {
      '1.5' : ['2#12,1#12G.', '1/2"', '30/2', '10A', 'FRN-R'],
      '3' : ['2#12,1#12G.', '1/2"','30/2', '20A', 'FRN-R'],
      '5' : ['2#10,1#10G.', '3/4"', '30/2', '30A', 'FRN-R'],
      '7.5' : ['2#6,1#10G.', '3/4"', '60/2', '45A', 'FRN-R'],
  }
  
  table_electric_heaters_208_3_KW = {
      '1.5' : ['3#12,1#12G.', '1/2"', '30/3', '6A', 'FRN-R'],
      '3' : ['3#12,1#12G.', '1/2"','30/3', '15A', 'FRN-R'],
      '5' : ['3#12,1#12G.', '1/2"','30/3', '20A', 'FRN-R'],
      '7.5' : ['3#10,1#10G.', '3/4"','30/3', '30A', 'FRN-R'],
      '10' : ['3#8,1#10G.', '3/4"','60/3', '40A', 'FRN-R'],
      '15' : ['3#6,1#10G.', '1"','60/3', '60A', 'FRN-R'],
      '20' : ['3#4,1#8G.', '1-1/2"','100/3', '70A', 'FRN-R'],
  }
  
  table_electric_heaters_277_1_KW = {
      '1.5' : ['2#12,1#12G.', '1/2"', '$SPST', '---'],
      '3' : ['2#12,1#12G.', '1/2"', '$SPST', '---'],
      '5' : ['2#10,1#10G.', '3/4"','30/2', '25A', 'FRS-R'],
      '7.5' : ['2#8,1#10G.', '3/4"','30/2', '40A', 'FRS-R'],
  }
  
  table_electric_heaters_480_3_KW = {
      '5' : ['3#12,1#12G.', '1/2"', '30/3', '9A', 'FRS-R'],
      '7.5' : ['3#12,1#12G.', '1/2"', '30/3', '15A', 'FRS-R'],
      '10' : ['3#12,1#12G.', '1/2"','30/3', '20A', 'FRS-R'],
      '15' : ['3#10,1#10G.', '3/4"','30/3', '30A', 'FRS-R'],
      '20' : ['3#10,1#10G.', '3/4"','30/3', '30A', 'FRS-R'],
      '25' : ['3#8,1#10G.', '1"','60/3', '40A', 'FRS-R'],
      '30' : ['3#6,1#10G.', '1"','60/3', '50A', 'FRS-R'],
  }
  
  table_120_1_MOCP = {
      '5' : ['2#12,1#12G.', '1/2"', '$T0', '5A', 'FRN-R'],
      '10' : ['2#12,1#12G.', '1/2"', '$T0', '10A', 'FRN-R'],
      '15' : ['2#12,1#12G.', '1/2"', '$T0', '15A', 'FRN-R'],
      '20' : ['2#12,1#12G.', '1/2"', '$T0', '20A', 'FRN-R'],
      '25' : ['2#10,1#10G.', '3/4"', '30/2', '25A', 'FRN-R'],
      '30' : ['2#10,1#10G.', '3/4"', '30/2', '30A', 'FRN-R'],
      '35' : ['2#8,1#10G.', '3/4"', '60/2', '35A', 'FRN-R'],
      '40' : ['2#8,1#10G.', '3/4"', '60/2', '40A', 'FRN-R'],
      '45' : ['2#6,1#10G.', '3/4"', '60/2', '45A', 'FRN-R'],
      '50' : ['2#6,1#10G.', '3/4"', '60/2', '50A', 'FRN-R'],
      '60' : ['2#4,1#8G.', '1"', '60/2', '60A', 'FRN-R'],
      '70' : ['2#4,1#8G.', '1"', '60/2', '70A', 'FRN-R'],
      '80' : ['2#3,1#8G.', '1"', '100/2', '80A', 'FRN-R'],
      '90' : ['2#2,1#8G.', '1-1/4"', '100/2', '90A', 'FRN-R'],
      '100' : ['2#1,1#6G.', '1-1/4"', '100/2', '100A', 'FRN-R'],
      '110' : ['2#1,1#6G.', '1-1/4"', '200/2', '110A', 'FRN-R'],
      '125' : ['2-1/0,#6G.', '1-1/2"C', '200/2', '125A', 'FRN-R'],
      '150' : ['2-1/0,#6G.', '1-1/2"C', '200/2', '150A', 'FRN-R'],
      'something else' : ['blah']
  }
  
  table_120_1_MCA = {
      '5' : ['2#12,1#12G.', '1/2"', '$T0', '---'],
      '10' : ['2#12,1#12G.', '1/2"', '$T0', '---'],
      '15' : ['2#12,1#12G.', '1/2"', '$T0', '---'],
      '20' : ['2#12,1#12G.', '1/2"', '$T0', '20A', 'FRN-R'],
      '25' : ['2#10,1#10G.', '3/4"', '30/2', '25A', 'FRN-R'],
      '30' : ['2#10,1#10G.', '3/4"', '30/2', '30A', 'FRN-R'],
      '35' : ['2#8,1#10G.', '3/4"', '60/2', '35A', 'FRN-R'],
      '40' : ['2#8,1#10G.', '3/4"', '60/2', '40A', 'FRN-R'],
      '45' : ['2#6,1#10G.', '3/4"', '60/2', '45A', 'FRN-R'],
      '50' : ['2#6,1#10G.', '3/4"', '60/2', '50A', 'FRN-R'],
      '60' : ['2#4,1#8G.', '1"', '60/2', '60A', 'FRN-R'],
      '70' : ['2#4,1#8G.', '1"', '60/2', '70A', 'FRN-R'],
      '80' : ['2#3,1#8G.', '1"', '100/2', '80A', 'FRN-R'],
      '90' : ['2#2,1#8G.', '1-1/4"', '100/2', '90A', 'FRN-R'],
      '100' : ['2#1,1#6G.', '1-1/4"', '100/2', '100A', 'FRN-R'],
      '110' : ['2#1,1#6G.', '1-1/4"', '200/2', '110A', 'FRN-R'],
      '125' : ['2-1/0,#6G.', '1-1/2"C', '200/2', '125A', 'FRN-R'],
      '150' : ['2-1/0,#6G.', '1-1/2"C', '200/2', '150A', 'FRN-R'],
      'something else' : ['blah']
  }
  
  table_120_3_MCA = {
      '5' : ['3#12,#12G.', '1/2"', '30/3', '5A', '---'],
      '10' : ['3#12,#12G.', '1/2"', '30/3', '10A', '---'],
      '15' : ['3#12,#12G.', '1/2"', '30/3', '15A', '---'],
      '20' : ['3#12,#12G.', '1/2"', '30/3', '20A', 'FRN-R'],
      '25' : ['3#10,#10G.', '3/4"', '30/2', '25A', 'FRN-R'],
      '30' : ['3#10,#10G.', '3/4"', '30/2', '30A', 'FRN-R'],
      '35' : ['3#8,#10G.', '3/4"', '60/2', '35A', 'FRN-R'],
      '40' : ['3#8,#10G.', '3/4"', '60/2', '40A', 'FRN-R'],
      '45' : ['3#6,#10G.', '1"', '60/2', '45A', 'FRN-R'],
      '50' : ['3#6,#10G.', '1"', '60/2', '50A', 'FRN-R'],
      '60' : ['3#4,#8G.', '1"', '60/2', '60A', 'FRN-R'],
      '70' : ['3#4,#8G.', '1"', '60/2', '70A', 'FRN-R'],
      '80' : ['3#3,#8G.', '1"', '100/2', '80A', 'FRN-R'],
      '90' : ['3#2,#8G.', '1-1/4"', '100/2', '90A', 'FRN-R'],
      '100' : ['3#1,#6G.', '1-1/2"', '100/2', '100A', 'FRN-R'],
      '110' : ['3#1,#6G.', '1-1/2"', '200/2', '110A', 'FRN-R'],
      '125' : ['3-1/0, #6G', '2"', '200/2', '125A', 'FRN-R'],
      '150' : ['3-1/0, #6G', '2"', '200/2', '150A', 'FRN-R'],
      '175' : ['3-2/0, #6G', '2"', '200/2', '175A', 'FRN-R'],
      '200' : ['3-3/0, #6G', '2"', '200/2', '200A', 'FRN-R'],
      '225' : ['3-4/0, #4G', '2"', '200/2', '225A', 'FRN-R'],
      '250' : ['3-250kCMIL, #4G', '3"', '200/2', '250A', 'FRN-R'],
      '300' : ['3-350kCMIL, #3G', '3"', '200/2', '300A', 'FRN-R'],
      '350' : ['2[3-2/0, #3G]', '(2) 2"', '200/2', '350A', 'FRN-R'],
      '400' : ['2[3-2/0, #3G]', '(2) 2"', '200/2', '400A', 'FRN-R'],
      '450' : ['2[3-250kCMIL, #2G]', '(2) 3"', '200/2', '450A', 'FRN-R'],
      '500' : ['2[3-250kCMIL, #2G]', '(2) 3"', '200/2', '500A', 'FRN-R'],
      '600' : ['2[3-350kCMIL, #1G]', '(2) 3"', '200/2', '600A', 'FRN-R'],
      'something else' : ['blah']
  }
  
  table_120_1_HP = {
      '1/6' : ['2#12,1#12G.', '1/2"', '$T0', '---'],
      '1/4' : ['2#12,1#12G.', '1/2"', '$T0', '---'],
      '1/3' : ['2#12,1#12G.', '1/2"', '$T0', '---'],
      '1/2' : ['2#12,1#12G.', '1/2"', '$T0', '---'],
      '3/4' : ['2#10,1#10G.', '3/4"', '30/2', '30A', 'FRN-R'],
      '1' : ['2#10,1#10G.', '3/4"', '30/2', '30A', 'FRN-R'],
      '1-1/2' : ['2#8,1#10G.', '3/4"', '60/2', '40A', 'FRN-R'],
      '2' : ['2#8,1#10G.', '3/4"', '60/2', '40A', 'FRN-R'],
      '3' : ['2#6,1#10G.', '3/4"', '60/2', '60A', 'FRN-R'],
      '5' : ['2#4,1#8G.', '1"', '100/2', '100A', 'FRN-R'],
      '7-1/2' : ['2#3,1#6G.', '1"', '200/2', '150A', 'FRN-R'],
      '10' : ['2#1,1#6G.', '1-1/4"', '200/2', '175A', 'FRN-R'],
  }
  
  table_208_1_HP = {
      '1/6' : ['2#12,1#12G.', '1/2"', '30/2', '6A', 'FRN-R'],
      '1/4' : ['2#12,1#12G.', '1/2"', '30/2', '6A', 'FRN-R'],
      '1/3' : ['2#12,1#12G.', '1/2"', '30/2', '10A', 'FRN-R'],
      '1/2' : ['2#12,1#12G.', '1/2"', '30/2', '10A', 'FRN-R'],
      '3/4' : ['2#12,1#12G.', '1/2"', '30/2', '15A', 'FRN-R'],
      '1' : ['2#12,1#12G.', '1/2"', '30/2', '15A', 'FRN-R'],
      '1-1/2' : ['2#12,1#12G.', '1/2"', '30/2', '20A', 'FRN-R'],
      '2' : ['2#10,1#10G.', '3/4"', '30/2', '25A', 'FRN-R'],
      '3' : ['2#10,1#10G.', '3/4"', '30/2', '30A', 'FRN-R'],
      '5' : ['2#6,1#10G.', '3/4"', '60/2', '60A', 'FRN-R'],
      '7-1/2' : ['2#4,1#8G.', '1"', '100/2', '80A', 'FRN-R'],
      '10' : ['2#4,1#8G.', '1"', '100/2', '100A', 'FRN-R'],
  }
  
  table_208_1_AMPS = {
      '1/6' : ['2#12,1#12G.', '1/2"', '30/2', '6A', 'FRN-R'],
      '1/4' : ['2#12,1#12G.', '1/2"', '30/2', '6A', 'FRN-R'],
      '1/3' : ['2#12,1#12G.', '1/2"', '30/2', '10A', 'FRN-R'],
      '1/2' : ['2#12,1#12G.', '1/2"', '30/2', '10A', 'FRN-R'],
      '3/4' : ['2#12,1#12G.', '1/2"', '30/2', '15A', 'FRN-R'],
      '1' : ['2#12,1#12G.', '1/2"', '30/2', '15A', 'FRN-R'],
      '1-1/2' : ['2#12,1#12G.', '1/2"', '30/2', '20A', 'FRN-R'],
      '2' : ['2#10,1#10G.', '3/4"', '30/2', '25A', 'FRN-R'],
      '3' : ['2#10,1#10G.', '3/4"', '30/2', '30A', 'FRN-R'],
      '5' : ['2#6,1#10G.', '3/4"', '60/2', '60A', 'FRN-R'],
      '7-1/2' : ['2#4,1#8G.', '1"', '100/2', '80A', 'FRN-R'],
      '10' : ['2#4,1#8G.', '1"', '100/2', '100A', 'FRN-R'],
      
  }
  
  table_208_3_HP = {
      '1/2' : ['3#12,1#12G.', '1/2"', '30/3', '6A', 'FRN-R'],
      '3/4' : ['3#12,1#12G.', '1/2"', '30/3', '6A', 'FRN-R'],
      '1' : ['3#12,1#12G.', '1/2"', '30/3', '10A', 'FRN-R'],
      '1-1/2' : ['3#12,1#12G.', '1/2"', '30/3', '15A', 'FRN-R'],
      '2' : ['3#12,1#12G.', '1/2"', '30/3', '15A', 'FRN-R'],
      '3' : ['3#12,1#12G.', '1/2"', '30/3', '20A', 'FRN-R'],
      '5' : ['3#10,1#10G.', '3/4"', '30/3', '30A', 'FRN-R'],
      '7-1/2' : ['3#8,1#10G.', '3/4"', '60/3', '40A', 'FRN-R'],
      '10' : ['3#8,1#10G.', '3/4"', '60/3', '50A', 'FRN-R'],
      '15' : ['3#4,1#8G.', '1"', '100/3', '80A', 'FRN-R'],
      '20' : ['3#3,1#8G.', '1-1/4"', '100/3', '100A', 'FRN-R'],
      '25' : ['3#2,1#6G.', '1-1/4"', '200/3', '125A', 'FRN-R'],
      '30' : ['3#1,1#6G.', '1-1/4"', '200/3', '150A', 'FRN-R'],
      '40' : ['3#1/0,1#6G.', '1-1/2"', '100/2', '200A', 'FRN-R'],
      '50' : ['3#3/0,1#4G.', '2"', '400/3', '250A', 'FRN-R'],
      '60' : ['3#4/0,1#4G.', '2"', '400/3', '300A', 'FRN-R'],
      '75' : ['3#250kcmil,1#3G.', '2-1/2"', '400/3', '350A', 'FRN-R'],
      '100' : ['3#500kcmil,1#2G.', '3"', '600/3', '500A', 'FRN-R'],
      
  }
  
  table_480_3_HP = {
      '1/2' : ['3#12,1#12G.', '1/2"', '30/3', '3A', 'FRS-R'],
      '3/4' : ['3#12,1#12G.', '1/2"', '30/3', '3A', 'FRS-R'],
      '1' : ['3#12,1#12G.', '1/2"', '30/3', '6A', 'FRS-R'],
      '1-1/2' : ['3#12,1#12G.', '1/2"', '30/3', '6A', 'FRS-R'],
      '2' : ['3#12,1#12G.', '1/2"', '30/3', '6A', 'FRS-R'],
      '3' : ['3#12,1#12G.', '1/2"', '30/3', '10A', 'FRS-R'],
      '5' : ['3#12,1#12G.', '1/2"', '30/3', '15A', 'FRS-R'],
      '7-1/2' : ['3#12,1#12G.', '1/2"', '30/3', '20A', 'FRS-R'],
      '10' : ['3#12,1#12G.', '1/2"', '30/3', '20A', 'FRS-R'],
      '15' : ['3#10,1#10G.', '3/4"', '30/3', '30A', 'FRS-R'],
      '20' : ['3#8,1#10G.', '3/4"', '60/3', '40A', 'FRS-R'],
      '25' : ['3#6,1#10G.', '1"', '60/3', '50A', 'FRS-R'],
      '30' : ['3#6,1#10G.', '1"', '60/3', '60A', 'FRS-R'],
      '40' : ['3#6,1#8G.', '1"', '100/2', '80A', 'FRS-R'],
      '50' : ['3#3,1#8G.', '1-1/4"', '100/3', '100A', 'FRS-R'],
      '60' : ['3#1,1#6G.', '1-1/4"', '200/3', '125A', 'FRS-R'],
      '75' : ['3#1/0,1#6G.', '1-1/2"', '200/3', '150A', 'FRS-R'],
      '100' : ['3#3/0,1#6G.', '2"', '200/3', '175A', 'FRS-R'],
      '125' : ['3#4/0,1#4G.', '2-1/2"', '400/3', '250A', 'FRS-R'],
      '150' : ['3#250kcmil,1#4G.', '2-1/2"', '400/3', '300A', 'FRS-R'],
      '200' : ['3#400kcmil,1#3G.', '3"', '400/3', '400A', 'FRS-R'],
      
  }
  
  def EQ_120_1_MCA(number): 
      
      hp = ''
      print(f'number = {number}')
      match number:
          case num if 0 < num <= 5:
              hp = '5'
          case num if 5 < num <= 10:
              hp = '10'
          case num if 10 < num <= 15:
              hp = '15'
          case num if 15 < num <= 20:
              hp = '20'
          case num if 20 < num <= 25:
              hp = '25'
          case num if 25 < num <= 30:
              hp = '30'
          case num if 30 < num <= 35:
              hp = '35'
          case num if 35 < num <= 40:
              hp = '40'
          case num if 40 < num <= 45:
              hp = '45'
          case num if 45 < num <= 50:
              hp = '50'
          case num if 50 < num <= 60:
              hp = '60'
          case num if 60 < num <= 70:
              hp = '70'
          case num if 70 < num <= 80:
              hp = '80'
          case num if 80 < num <= 90:
              hp = '90'
          case num if 90 < num <= 100:
              hp = '100'
          case num if 100 < num <= 110:
              hp = '110'
          case num if 110 < num <= 125:
              hp = '125'
          case num if 125 < num <= 150:
              hp = '150'
          case _:
              hp = 'something else'
      print(f'hp = {hp}')
      return hp
  
  def EQ_120_3_MCA(number): 
      
      hp = ''
      print(f'number = {number}')
      match number:
          case num if 0 < num <= 5:
              hp = '5'
          case num if 5 < num <= 10:
              hp = '10'
          case num if 10 < num <= 15:
              hp = '15'
          case num if 15 < num <= 20:
              hp = '20'
          case num if 20 < num <= 25:
              hp = '25'
          case num if 25 < num <= 30:
              hp = '30'
          case num if 30 < num <= 35:
              hp = '35'
          case num if 35 < num <= 40:
              hp = '40'
          case num if 40 < num <= 45:
              hp = '45'
          case num if 45 < num <= 50:
              hp = '50'
          case num if 50 < num <= 60:
              hp = '60'
          case num if 60 < num <= 70:
              hp = '70'
          case num if 70 < num <= 80:
              hp = '80'
          case num if 80 < num <= 90:
              hp = '90'
          case num if 90 < num <= 100:
              hp = '100'
          case num if 100 < num <= 110:
              hp = '110'
          case num if 110 < num <= 125:
              hp = '125'
          case num if 125 < num <= 150:
              hp = '150'
          case num if 150 < num <= 175:
              hp = '175'
          case num if 175 < num <= 200:
              hp = '200'
          case num if 200 < num <= 225:
              hp = '225'
          case num if 225 < num <= 250:
              hp = '250'
          case num if 250 < num <= 300:
              hp = '300'
          case num if 300 < num <= 350:
              hp = '350'
          case num if 350 < num <= 400:
              hp = '400'
          case num if 400 < num <= 450:
              hp = '450'
          case num if 450 < num <= 500:
              hp = '500'
          case num if 500 < num <= 600:
              hp = '600'    
          case _:
              hp = 'something else'
      print(f'hp = {hp}')
      return hp
  
  # def EQ_120_1_AMPS(number): 
      
  #     hp = ''
  
  #     print(f'number = {number}')
  #     match number:
  #         case num if 0 < num <= 4.4:
  #             hp = '1/6'
  #         case num if 4.4 < num <= 5.8:
  #             hp = '1/4'
  #         case num if 5.8 < num <= 7.2:
  #             hp = '1/3'
  #         case num if 7.2 < num <= 9.8:
  #             hp = '1/2'
  #         case num if 9.8 < num <= 13.8:
  #             hp = '3/4'
  #         case num if 13.8 < num <= 16:
  #             hp = '1'
  #         case num if 16 < num <= 20:
  #             hp = '1-1/2'
  #         case num if 20 < num <= 24:
  #             hp = '2'
  #         case num if 24 < num <= 34:
  #             hp = '3'
  #         case num if 34 < num <= 56:
  #             hp = '5'
  #         case num if 56 < num <= 80:
  #             hp = '7-1/2'
  #         case num if 80 < num <= 100:
  #             hp = '10'
  #         case _:
  #             hp = 'not found'
  #     print(f'hp = {hp}')
  #     return hp
  
  def EQ_120_1_KW(number): 
      
      hp = ''
      print(f'number = {number}')
      match number:
          case num if 0 < num <= 0.5:
              hp = '0.5'
          case num if 0.5 < num <= 1.5:
              hp = '1.5'
          case _:
              hp = 'not found'
      print(f'hp = {hp}')
      return hp
  
  def EQ_208_1_KW(number): 
      
      hp = ''
      print(f'number = {number}')
      match number:
          case num if 0 < num <= 1.5:
              hp = '1.5'
          case num if 1.5 < num <= 3:
              hp = '3'
          case num if 3 < num <= 5:
              hp = '5'
          case num if 5 < num <= 7.5:
              hp = '7.5'
          case _:
              hp = 'not found'
      print(f'hp = {hp}')
      return hp
  
  def EQ_208_3_KW(number): 
      
      hp = ''
      print(f'number = {number}')
      match number:
          case num if 0 < num <= 1.5:
              hp = '1.5'
          case num if 1.5 < num <= 3:
              hp = '3'
          case num if 3 < num <= 5:
              hp = '5'
          case num if 5 < num <= 7.5:
              hp = '7.5'
          case num if 7.5 < num <= 10:
              hp = '10'
          case num if 10 < num <= 15:
              hp = '15'        
          case num if 15 < num <= 20:
              hp = '20'
          case _:
              hp = 'not found'
      print(f'hp = {hp}')
      return hp
  
  def EQ_277_1_KW(number): 
      
      hp = ''
      print(f'number = {number}')
      match number:
          case num if 0 < num <= 1.5:
              hp = '1.5'
          case num if 1.5 < num <= 3:
              hp = '3'
          case num if 3 < num <= 5:
              hp = '5'
          case num if 5 < num <= 7.5:
              hp = '7.5'
          case _:
              hp = 'not found'
      print(f'hp = {hp}')
      return hp
  
  def EQ_208_1_MCA(number): 
      
      hp = ''
      print(f'number = {number}')
      match number:
          case num if 0 < num <= 3:
              hp = '1/6'
          case num if 3 < num <= 4:
              hp = '1/4'
          case num if 4 < num <= 5:
              hp = '1/3'
          case num if 5 < num <= 6.75:
              hp = '1/2'
          case num if 6.75 < num <= 9.5:
              hp = '3/4'
          case _:
              hp = 'not found'
      print(f'hp = {hp}')
      return hp
  
  def EQ_208_3_MCA(number): 
      
      hp = ''
      print(f'number = {number}')
      match number:
          case num if 0 < num <= 3:
              hp = '1/6'
          case num if 3 < num <= 4:
              hp = '1/4'
          case num if 4 < num <= 5:
              hp = '1/3'
          case num if 5 < num <= 6.75:
              hp = '1/2'
          case num if 6.75 < num <= 9.5:
              hp = '3/4'
          case _:
              hp = 'not found'
      print(f'hp = {hp}')
      return hp
  
  def EQ_208_1_AMPS(number): 
      
      hp = ''
      print(f'number = {number}')
      match number:
          case num if 0 < num <= 3:
              hp = '1/6'
          case num if 3 < num <= 4:
              hp = '1/4'
          case num if 4 < num <= 5:
              hp = '1/3'
          case num if 5 < num <= 6.75:
              hp = '1/2'
          case num if 6.75 < num <= 9.5:
              hp = '3/4'
          case _:
              
              hp = f'{load} not found'
      print(f'hp = {hp}')
      return hp
  
  def EQ_480_3_MCA(number): 
      
      hp = ''
      print(f'number = {number}')
      match number:
          case num if 0 < num <= 3:
              hp = '1/6'
          case num if 3 < num <= 4:
              hp = '1/4'
          case num if 4 < num <= 5:
              hp = '1/3'
          case num if 5 < num <= 6.75:
              hp = '1/2'
          case num if 6.75 < num <= 9.5:
              hp = '3/4'
          case _:
              
              hp = f'{load} not found'
      print(f'hp = {hp}')
      return hp
  
  def volt_phase_check_HP(load, voltage, phase):
      
      size = []
          
      if voltage == 120 and phase == 1:
          size = table_120_1_HP[load]
      if voltage == 208 and phase == 1:
          size = table_208_1_HP[load]
      if voltage == 208 and phase == 3:
          size = table_208_3_HP[load]
      if voltage == 480 and phase == 3:
          size = table_480_3_HP[load]        
      return size
  
  def volt_phase_check_MCA(load, voltage, phase):
      
      size = []
          
      if voltage == 120 and phase == 1:
          val = EQ_120_1_MCA(load)
          size = table_120_1_MCA[val]
      if voltage == 120 and phase == 3:
          val = EQ_120_3_MCA(load)
          size = table_120_3_MCA[val]    
      if voltage == 208 and phase == 1:
          val = EQ_120_1_MCA(load)
          size = table_120_1_MCA[val]
      if voltage == 208 and phase == 3:
          val = EQ_120_3_MCA(load)
          size = table_120_3_MCA[val]        
      if voltage == 480 and phase == 3:
          val = EQ_120_3_MCA(load)
          size = table_120_3_MCA[val] 
      return size
  
  def volt_phase_check_AMPS(load, voltage, phase):
      
      size = []
      
      MCA = (load * 1.25)
      print('MCA = ' + str(MCA))
      if voltage == 120 and phase == 1:
          val = EQ_120_1_MCA(MCA)
          size = table_120_1_MCA[val]
      if voltage == 120 and phase == 3:
          val = EQ_120_3_MCA(MCA)
          size = table_120_3_MCA[val]
      if voltage == 208 and phase == 1:
          val = EQ_120_1_MCA(MCA)
          size = table_120_1_MCA[val]
      if voltage == 208 and phase == 3:
          val = EQ_120_3_MCA(MCA)
          size = table_120_3_MCA[val]       
      if voltage == 480 and phase == 3:
          val = EQ_120_3_MCA(MCA)
          size = table_120_3_MCA[val]
  
      print(val)        
      return size
  
  def volt_phase_check_WATTS(load, voltage, phase):
      
      size = []
      
      
      KW = round(float(load / 1000) , 3)
      
      print(KW)
      if voltage == 120 and phase == 1:
          val = EQ_120_1_KW(KW)
          size = table_electric_heaters_120_1_KW[val]
      if voltage == 208 and phase == 1:
          val = EQ_208_1_KW(KW)
          size = table_electric_heaters_208_1_KW[val]
      if voltage == 208 and phase == 3:
          val = EQ_208_3_KW(KW)
          size = table_electric_heaters_208_3_KW[val]
      if voltage == 277 and phase == 1:
          val = EQ_277_1_KW(KW)
          size = table_electric_heaters_277_1_KW[val]   
      
      return size
  
  def volt_phase_check_MOP(load, voltage, phase):
      
      size = []
      
      print(load)
      
      if voltage == 120 and phase == 1:
          size = table_120_1_MOCP[load]
      if voltage == 120 and phase == 3:
          size = table_120_3_MCA[load]
      if voltage == 208 and phase == 1:
          size = table_120_1_MOCP[load]
      if voltage == 208 and phase == 3:
          size = table_120_3_MCA[load]
      return size
  
  def volt_phase_check_KW(load, voltage, phase):
      
      size = []
          
      if voltage == 120 and phase == 1:
          val = EQ_120_1_KW(load)
          size = table_electric_heaters_120_1_KW[val]
      if voltage == 208 and phase == 1:
          val = EQ_208_1_KW(load)
          size = table_electric_heaters_208_1_KW[val]
      if voltage == 208 and phase == 3:
          val = EQ_208_3_KW(load)
          size = table_electric_heaters_208_3_KW[val]
      if voltage == 277 and phase == 1:
          val = EQ_277_1_KW(load)
          size = table_electric_heaters_277_1_KW[val]        
      return size
  
  def parse_load_value_MOP(load_value):
      
      trimmed_load_value = load_value.strip()
      
      load, load_type = trimmed_load_value.split(' ')
      
      print(f'load = {load} : type = {load_type}')
      
      return str(load), load_type
  
  def parse_load_value(load_value):
      
      trimmed_load_value = load_value.strip()
      
      load, load_type = trimmed_load_value.split(' ')
      
      print(f'load = {load} : type = {load_type}')
      
      return float(load), load_type
  
  def parse_load_value_HP(load_value):
      
      load, load_type = load_value.split(' ')
      
      print(f'load = {load} : type = {load_type}')
      
      return str(load), load_type
  
  for cell in SH['C']:
  
      load = cell.value
      voltage = SH[f'D{cell.row}'].value
      phase = SH[f'E{cell.row}'].value
  
      if cell.value is None:
          continue
      if cell.value == 'LOAD':
          continue
      if cell.value == 'VOLTS':
          continue
      if cell.value == '%%C':
          continue
      if cell.value == 'HP,KW,FLA':
          continue
      if 'HP' in cell.value:
        try:
          load, load_type = parse_load_value_HP(cell.value)
          voltage = SH[f'D{cell.row}'].value
          phase = SH[f'E{cell.row}'].value
  
          if load_type == 'HP':
              for i, value in enumerate(volt_phase_check_HP(load, voltage, phase)):
                  SH.cell(column=i+6, row=cell.row, value=value)
              continue
        except:
            pass
  
      if 'MOP' or 'MOCP' in cell.value:
        try: 
          load, load_type = parse_load_value_MOP(cell.value)
          voltage = SH[f'D{cell.row}'].value
          phase = SH[f'E{cell.row}'].value
        
          if load_type == 'MOP':
              for i, value in enumerate(volt_phase_check_MOP(load, voltage, phase)):
                  SH.cell(column=i+6, row=cell.row, value=value)
              continue
          if load_type == 'MOCP':
              for i, value in enumerate(volt_phase_check_MOP(load, voltage, phase)):
                  SH.cell(column=i+6, row=cell.row, value=value)
              continue
        except:
            pass      
             
      if cell.value is not None:
        try:  
          load, load_type = parse_load_value(cell.value)
          voltage = SH[f'D{cell.row}'].value
          phase = SH[f'E{cell.row}'].value
  
          if voltage == None:
              continue
          if phase == None:
              continue
          if load_type == None:
              continue
          if load_type == 'MCA':
              for i, value in enumerate(volt_phase_check_MCA(load, voltage, phase)):
                  SH.cell(column=i+6, row=cell.row, value=value)
                  #SH.move_range("J:{cell.row}", rows=2, cols=3)
          if load_type == 'AMPS' or 'FLA':
              for i, value in enumerate(volt_phase_check_AMPS(load, voltage, phase)):
                  SH.cell(column=i+6, row=cell.row, value=value)
          if load_type == 'KW':
              for i, value in enumerate(volt_phase_check_KW(load, voltage, phase)):
                  SH.cell(column=i+6, row=cell.row, value=value)
          if load_type == 'W':
              for i, value in enumerate(volt_phase_check_WATTS(load, voltage, phase)):
                  SH.cell(column=i+6, row=cell.row, value=value) 
            
              print(f'Voltage = {voltage} : Phase = {phase}')
              print(volt_phase_check_MCA(load, voltage, phase))
                
          elif load_type != 'MCA' or 'AMPS' or 'FLA' or 'KW' or 'W':
              continue        

          else:
              continue
        except:
            pass  
  # for cell in SH['J']:
  #   if cell.value is None:
  #       continue
  #   if cell.value == 'PANEL':
  #       continue
  #   if cell.value == 'FRN-R' or 'FRS-R' or 'RK1' or '---':
  #       SH.move_range(f'J{cell.row}:J{cell.row}', rows=+ 1, cols= -1, translate=True)    
  
  # for cell in SH['I']:
  #     if cell.value == 'FUSE':
  #         continue
  #     if cell.value is None:
  #         continue
  #     if 'A' in cell.value:
  #         continue
  #     if cell.value == '---':
  #         continue
  #     if 'FRN-R' or 'FRS-R' or 'RK1' in cell.value:
  #         cell.border = copy(SH["J5"].border) 
  #         cell.font = copy(SH["J2"].font)
  #         cell.alignment = copy(SH["J2"].alignment)   
  WB.save(uploaded_file)  
 
if uploaded_file is not None:
  file_details = {"FileName":uploaded_file.name,"FileType":uploaded_file.type,"FileSize":uploaded_file.size}
  file_name, file_type = uploaded_file.name.split('.')
  st.write(file_details)
  if st.button('Do it for me!'):
    Automation_xl(uploaded_file)
    buffer = io.BytesIO()
    WB.save(buffer)
    df = pd.DataFrame(Automation_xl(uploaded_file))
    st.dataframe(display_xl(), width = 1500)
    st.write("Are you happy with these results?")   
    buffer = io.BytesIO()
    WB.save(buffer)    
    st.download_button(
        label = 'Hellz Yeah! Click to Download 📥', 
        data = buffer, 
        file_name = uploaded_file.name, 
        mime= uploaded_file.type
        )
    
    st.text('If nothing is produced in desired cell range, please double check the formatting of LOADS/VOLTS/PHASE. (see reference table above)')  
    st.text('If results are produced but are incorrect, please contact Owen for changes to be made.')   
    st.text('If you managed to break the code, congrats! Please send screenshot of the broken string as well as Equipment Schedule to Owen')

# SIDEBAR

# st.sidebar.header("There Will Be More Options Here Soon:")

hide_st_style="""
            <style>
            #MainMenu {visibility:hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            </style>
            """
 
st.markdown(hide_st_style, unsafe_allow_html=True)
 

